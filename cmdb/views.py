import os, stat, time
import json
import paramiko
from concurrent.futures import ThreadPoolExecutor
from datetime import datetime
from django.shortcuts import render, redirect, get_object_or_404
from django.contrib.auth.decorators import login_required
from django.contrib import messages
from django.db.models import Q
from django.http import JsonResponse, FileResponse, HttpResponse
from django.utils.encoding import escape_uri_path
from django.views.decorators.csrf import csrf_exempt
from .models import HighRiskAudit

import openpyxl
from .models import SSLCertificate
from .tasks import check_ssl_certificates_task
from system.models import SystemConfig
try:
    from aliyunsdkcore.client import AcsClient
    from aliyunsdkecs.request.v20140526 import DescribeInstancesRequest

    ALIYUN_READY = True
except ImportError:
    ALIYUN_READY = False

# 2. 腾讯云
try:
    from tencentcloud.common import credential
    from tencentcloud.common.profile.client_profile import ClientProfile
    from tencentcloud.common.profile.http_profile import HttpProfile
    from tencentcloud.cvm.v20170312 import cvm_client, models as cvm_models

    TENCENT_READY = True
except ImportError:
    TENCENT_READY = False

# 引入本应用模型
from .models import Server, ServerGroup, CloudAccount, ServerGroupAuth, TerminalLog, ServerMetric
from .forms import ServerForm, GroupForm, CloudAccountForm
from .agent_code import AGENT_SCRIPT_CONTENT
# 引入 AI
from ai_ops.models import AIModel
from ai_ops.utils import ask_ai
from kubernetes import client, config, stream

try:
    from k8s_manager.models import Cluster as K8sCluster
    from k8s_manager.utils import get_k8s_client  # 假设您封装了获取 client 的方法
except ImportError:
    # 兼容性处理，防止报错
    K8sCluster = None


# ========================================================
# 0. 核心工具函数：安全 SSH 连接
# ========================================================
def get_secure_ssh_client(server, timeout=10):
    """
    获取安全的 SSH 客户端连接
    功能：
    1. 优先加载系统 known_hosts
    2. 加载/创建项目专属 ssh_known_hosts
    3. 首次连接自动保存 HostKey，后续连接校验 (防止 MITM)
    """
    client = paramiko.SSHClient()

    # 1. 加载系统级 Host Keys (如 ~/.ssh/known_hosts)
    client.load_system_host_keys()

    # 2. 加载项目级 Host Keys
    # 获取项目根目录 (假设当前在 cmdb/views.py -> parent -> parent = base_dir)
    base_dir = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
    known_hosts_path = os.path.join(base_dir, 'ssh_known_hosts')

    try:
        client.load_host_keys(known_hosts_path)
    except IOError:
        # 如果文件不存在，先创建一个空的
        with open(known_hosts_path, 'w') as f:
            pass

    # 3. 设置策略：首次自动添加并保存，后续变更会报错 (Reject)
    # 注意：AutoAddPolicy 默认只添加不保存，需要手动 save_host_keys
    client.set_missing_host_key_policy(paramiko.AutoAddPolicy())

    # 4. 建立连接
    try:
        client.connect(
            server.ip_address,
            port=server.port,
            username=server.username,
            password=server.password,
            timeout=timeout
        )

        # 5. 连接成功后，持久化保存 Host Key 到文件
        # 这样下一次连接时，load_host_keys 就会读取到，从而实现校验
        client.save_host_keys(known_hosts_path)

        return client
    except Exception as e:
        # 连接失败时确保清理资源
        client.close()
        raise e


# ========================================================
# 1. 服务器管理 (列表 & CRUD)
# ========================================================
@login_required
def server_list(request):
    user = request.user

    # 权限过滤
    if user.is_superuser:
        servers = Server.objects.all().order_by('-id')
        root_groups = ServerGroup.objects.filter(parent__isnull=True)
        allowed_group_ids = []
    else:
        my_role_ids = user.groups.values_list('id', flat=True)
        allowed_group_ids = list(
            ServerGroupAuth.objects.filter(role_id__in=my_role_ids).values_list('server_group_id', flat=True))
        servers = Server.objects.filter(group_id__in=allowed_group_ids).order_by('-id')
        root_groups = ServerGroup.objects.filter(id__in=allowed_group_ids)

    # 筛选
    group_id = request.GET.get('group')
    current_group_id = None
    if group_id:
        try:
            current_group_id = int(group_id)
            if not user.is_superuser and current_group_id not in allowed_group_ids:
                pass
            else:
                servers = servers.filter(group_id=current_group_id)
        except ValueError:
            pass

    # 搜索
    search_kw = request.GET.get('q', '').strip()
    if search_kw:
        servers = servers.filter(Q(hostname__icontains=search_kw) | Q(ip_address__icontains=search_kw))

    cloud_accounts = CloudAccount.objects.all() if user.is_superuser else []

    context = {
        'page_title': '主机列表',
        'root_groups': root_groups,
        'servers': servers,
        'current_group_id': current_group_id,
        'search_kw': search_kw,
        'cloud_accounts': cloud_accounts
    }
    return render(request, 'cmdb/server_list.html', context)


@login_required
def server_add(request):
    if not (request.user.is_superuser or request.user.has_perm('cmdb.add_server')):
        return redirect('server_list')
    if request.method == 'POST':
        form = ServerForm(request.POST)
        if form.is_valid():
            form.save()
            messages.success(request, "添加成功")
            return redirect('server_list')
    else:
        form = ServerForm()
    return render(request, 'cmdb/server_form.html', {'form': form, 'page_title': '添加主机'})


@login_required
def server_edit(request, pk):
    if not request.user.is_superuser: return redirect('server_list')
    server = get_object_or_404(Server, pk=pk)

    if request.method == 'POST':
        form = ServerForm(request.POST, instance=server)
        if form.is_valid():
            obj = form.save(commit=False)

            new_password = form.cleaned_data.get('password')

            if not new_password:
                old_obj = Server.objects.get(pk=pk)
                obj.password = old_obj.password

            obj.save()
            messages.success(request, "修改成功")
            return redirect('server_list')
    else:
        form = ServerForm(instance=server)

    return render(request, 'cmdb/server_form.html', {'form': form, 'page_title': '编辑主机'})


@login_required
def server_delete(request, pk):
    if request.user.is_superuser:
        get_object_or_404(Server, pk=pk).delete()
    return redirect('server_list')


@login_required
def group_add(request):
    if request.user.is_superuser and request.method == 'POST':
        form = GroupForm(request.POST)
        if form.is_valid(): form.save()
    return redirect('server_list')


# ========================================================
# 2. 阿里云/腾讯云同步
# ========================================================
@login_required
def sync_aliyun(request):
    """
    同步云资源入口 (支持阿里云和腾讯云)
    注意：虽然路由名叫 sync_aliyun，但逻辑已扩展支持多云
    """
    if not request.user.is_superuser or request.method != 'POST':
        return redirect('server_list')

    account_id = request.POST.get('account_id')
    try:
        acc = CloudAccount.objects.get(id=account_id)

        # 1. 阿里云同步
        if acc.type == 'aliyun':
            if not ALIYUN_READY:
                messages.error(request, "阿里云 SDK 未安装 (aliyun-python-sdk-ecs)")
                return redirect('server_list')

            _sync_aliyun_ecs(request, acc)

        # 2. 腾讯云同步
        elif acc.type == 'tencent':
            if not TENCENT_READY:
                messages.error(request, "腾讯云 SDK 未安装 (tencentcloud-sdk-python)")
                return redirect('server_list')

            _sync_tencent_cvm(request, acc)

        else:
            messages.error(request, f"不支持的云厂商类型: {acc.type}")

    except Exception as e:
        messages.error(request, f"同步流程异常: {str(e)}")

    return redirect('server_list')


def _sync_aliyun_ecs(request, acc):
    """阿里云 ECS 同步逻辑"""
    try:
        # acc.secret_key 会被 django-fernet-fields 自动解密
        client = AcsClient(acc.access_key, acc.secret_key, acc.region)
        req = DescribeInstancesRequest.DescribeInstancesRequest()
        req.set_PageSize(100)

        # 发送请求
        resp_bytes = client.do_action_with_exception(req)
        resp = json.loads(resp_bytes)

        count = 0
        for item in resp.get('Instances', {}).get('Instance', []):
            # 获取 IP (优先公网)
            public_ips = item.get('PublicIpAddress', {}).get('IpAddress', [])
            private_ips = item.get('VpcAttributes', {}).get('PrivateIpAddress', {}).get('IpAddress', [])
            ip = public_ips[0] if public_ips else (private_ips[0] if private_ips else '')

            if not ip: continue

            # 状态映射
            status = 'Running' if item['Status'] == 'Running' else 'Stopped'

            Server.objects.update_or_create(
                instance_id=item['InstanceId'],
                defaults={
                    'hostname': item['InstanceName'],
                    'ip_address': ip,
                    'cpu_cores': item['Cpu'],
                    'memory_gb': int(item['Memory'] / 1024),  # MB -> GB
                    'os_name': item['OSName'],
                    'status': status,
                    'provider': 'aliyun'
                }
            )
            count += 1
        messages.success(request, f"阿里云同步成功: 更新了 {count} 台主机")

    except Exception as e:
        raise Exception(f"阿里云 API 调用失败: {str(e)}")


def _sync_tencent_cvm(request, acc):
    """腾讯云 CVM 同步逻辑"""
    try:
        # 实例化认证对象 (自动解密 SK)
        cred = credential.Credential(acc.access_key, acc.secret_key)

        # 实例化 HTTP 选项
        httpProfile = HttpProfile()
        httpProfile.endpoint = "cvm.tencentcloudapi.com"

        # 实例化 Client
        clientProfile = ClientProfile()
        clientProfile.httpProfile = httpProfile
        client = cvm_client.CvmClient(cred, acc.region, clientProfile)

        # 实例化请求对象
        req = cvm_models.DescribeInstancesRequest()
        req.Limit = 100  # 单次最大 100

        # 发送请求
        resp = client.DescribeInstances(req)

        count = 0
        for item in resp.InstanceSet:
            # 获取 IP
            ip = item.PublicIpAddresses[0] if item.PublicIpAddresses else (
                item.PrivateIpAddresses[0] if item.PrivateIpAddresses else '')
            if not ip: continue

            # 状态映射 (Tencent: RUNNING, STOPPED)
            status = 'Running' if item.InstanceState == 'RUNNING' else 'Stopped'

            Server.objects.update_or_create(
                instance_id=item.InstanceId,
                defaults={
                    'hostname': item.InstanceName,
                    'ip_address': ip,
                    'cpu_cores': item.CPU,
                    'memory_gb': int(item.Memory / 1024),  # MB -> GB
                    'os_name': item.OsName,
                    'status': status,
                    'provider': 'tencent'  # 确保 Server model 的 choices 里有这个
                }
            )
            count += 1

        messages.success(request, f"腾讯云同步成功: 更新了 {count} 台主机")

    except Exception as e:
        raise Exception(f"腾讯云 API 调用失败: {str(e)}")


# ========================================================
# 3. WebSSH & 文件
# ========================================================
@login_required
def webssh(request, server_id):
    server = get_object_or_404(Server, id=server_id)
    if not request.user.is_superuser:
        # 获取用户所有授权的分组ID
        my_role_ids = request.user.groups.values_list('id', flat=True)
        allowed_group_ids = ServerGroupAuth.objects.filter(role_id__in=my_role_ids).values_list('server_group_id',
                                                                                                flat=True)

        # 检查服务器是否在授权分组内
        if not server.group or server.group.id not in allowed_group_ids:
            return HttpResponse("权限不足 (Permission Denied)", status=403)
    return render(request, 'cmdb/webssh.html', {
        'server': server,
        'ai_models': AIModel.objects.all(),
    })


@login_required
@csrf_exempt
def server_file_upload(request, server_id):
    if request.method != 'POST': return JsonResponse({'status': False})
    server = get_object_or_404(Server, id=server_id)
    f = request.FILES.get('file_data')
    path = request.POST.get('remote_path', '/tmp')
    if not f: return JsonResponse({'status': False, 'msg': '无文件'})

    client = None
    try:
        # 使用安全连接
        client = get_secure_ssh_client(server)
        sftp = client.open_sftp()
        full = os.path.join(path, f.name).replace('\\', '/')
        sftp.putfo(f, full)
        sftp.close()
        return JsonResponse({'status': True, 'msg': '上传成功'})
    except Exception as e:
        return JsonResponse({'status': False, 'msg': str(e)})
    finally:
        if client: client.close()


@login_required
def server_file_download(request, server_id):
    path = request.GET.get('filepath')
    if not path:
        return HttpResponse("missing path", status=400)

    server = get_object_or_404(Server, id=server_id)

    client = None
    sftp = None

    try:
        # 1. 提前建立连接 (使用安全连接)
        client = get_secure_ssh_client(server)
        sftp = client.open_sftp()

        # 2. 获取文件大小 (关键步骤：用于前端显示进度条)
        try:
            file_stat = sftp.stat(path)
            file_size = file_stat.st_size
        except IOError:
            # 如果文件不存在或无法访问，尽早抛出错误
            raise Exception(f"Remote file not found or inaccessible: {path}")

        # 3. 定义闭包生成器：复用已建立的 sftp 连接
        # 注意：这里不需要再 connect，直接使用外部的 sftp 对象
        def file_iterator(chunk_size=65536):
            try:
                with sftp.open(path, 'rb') as f:
                    while True:
                        chunk = f.read(chunk_size)
                        if not chunk:
                            break
                        yield chunk
            finally:
                # 4. 确保下载完成或中断时关闭连接
                if sftp: sftp.close()
                if client: client.close()

        # 5. 构建响应
        filename = os.path.basename(path)
        response = FileResponse(file_iterator(), as_attachment=True)

        # === 核心优化：设置 Content-Length ===
        response['Content-Length'] = file_size
        response['Content-Disposition'] = f"attachment; filename*=utf-8''{escape_uri_path(filename)}"

        return response

    except Exception as e:
        # 如果在建立连接或获取大小时就出错了，需要在这里手动关闭
        # 因为生成器还没开始运行，不会触发上面的 finally
        if sftp: sftp.close()
        if client: client.close()
        return HttpResponse(f"Error initiating download: {str(e)}", status=500)


@login_required
def terminal_log_list(request):
    if not request.user.is_superuser: return redirect('dashboard')
    logs = TerminalLog.objects.select_related('user', 'server').all().order_by('-start_time')
    return render(request, 'cmdb/log_list.html', {'logs': logs, 'ai_models': AIModel.objects.all()})


@login_required
def terminal_log_detail(request, log_id):
    if not request.user.is_superuser: return redirect('dashboard')
    log = get_object_or_404(TerminalLog, id=log_id)
    records = []
    if log.log_file:
        try:
            with log.log_file.open('r') as f:
                for line in f:
                    try:
                        records.append(json.loads(line))
                    except:
                        pass
        except:
            pass
    return render(request, 'cmdb/log_detail.html', {'log': log, 'records_json': json.dumps(records)})


# ========================================================
# 4. Agent 管理 (批量安装 & 上报接口)
# ========================================================
@login_required
def agent_install(request):
    if not request.user.is_superuser:  # <--- 新增超级管理员检查
        return JsonResponse({'status': False, 'msg': '权限不足'})
    """批量 SSH 安装 Agent (增强版：含 Pip 检查与运行验证)"""
    if request.method != 'POST': return JsonResponse({'status': False})
    server_ids = request.POST.getlist('server_ids[]')
    servers = Server.objects.filter(id__in=server_ids)

    def install_one(server):
        log = []
        client = None
        try:
            if not server.username or not server.password:
                return f"[{server.hostname}] ❌ 无 SSH 账号密码"

            # 使用安全连接
            client = get_secure_ssh_client(server, timeout=10)

            # === 1. 检查 Pip 是否存在 ===
            log.append(f"[{server.hostname}] 检查环境...")
            stdin, stdout, stderr = client.exec_command("which pip3")
            if stdout.channel.recv_exit_status() != 0:
                return f"[{server.hostname}] ❌ 安装失败: 未检测到 pip3，请先在服务器安装 python3-pip"

            # === 2. 安装依赖 ===
            log.append(f"[{server.hostname}] 安装依赖(psutil)...")
            cmd = "pip3 install psutil -i https://pypi.tuna.tsinghua.edu.cn/simple --timeout 15"
            stdin, stdout, stderr = client.exec_command(cmd)
            exit_code = stdout.channel.recv_exit_status()
            if exit_code != 0:
                # 即使依赖安装失败，有时也可以尝试继续，或者直接报错。这里选择记录警告。
                err_msg = stderr.read().decode().strip()
                log.append(f"⚠️ pip警告/错误: {err_msg[:50]}...")

            # === 3. 写入脚本并启动 ===
            log.append(f"[{server.hostname}] 写入脚本并启动...")
            remote = "/opt/ops_agent.py"
            sftp = client.open_sftp()
            with sftp.open(remote, 'w') as f:
                f.write(AGENT_SCRIPT_CONTENT)
            sftp.close()

            # 停止旧进程并启动新进程
            client.exec_command("pkill -f ops_agent.py")
            # 使用 nohup 启动
            client.exec_command(f"nohup python3 {remote} > /dev/null 2>&1 &")

            # === 4. 验证 Agent 是否成功运行 ===
            time.sleep(2)  # 等待2秒，让程序有时间启动或报错退出
            # 使用 pgrep 检查进程 (-f 匹配完整命令行)
            stdin, stdout, stderr = client.exec_command("pgrep -f ops_agent.py")
            if stdout.channel.recv_exit_status() != 0:
                # 进程未找到，说明启动失败
                return f"[{server.hostname}] ❌ 启动失败: 进程在启动后立即退出，请检查服务器日志或依赖"

            # 更新数据库状态
            server.use_agent = True
            server.save()
            return f"[{server.hostname}] ✅ 部署成功 (PID: {stdout.read().decode().strip()})"

        except Exception as e:
            return f"[{server.hostname}] ❌ 异常: {str(e)}"
        finally:
            if client: client.close()

    with ThreadPoolExecutor(max_workers=10) as ex:
        futures = [ex.submit(install_one, s) for s in servers]
        results = [f.result() for f in futures]

    return JsonResponse({'status': True, 'log': "\n".join(results)})


@login_required
@csrf_exempt
def agent_uninstall(request):
    """批量卸载 Agent"""
    if request.method != 'POST': return JsonResponse({'status': False})
    server_ids = request.POST.getlist('server_ids[]')
    servers = Server.objects.filter(id__in=server_ids)

    def uninstall_one(server):
        client = None
        try:
            if not server.username or not server.password:
                return f"[{server.hostname}] ❌ 无 SSH 账号密码"

            # 使用安全连接
            client = get_secure_ssh_client(server, timeout=10)

            # 1. 停止进程
            client.exec_command("pkill -f ops_agent.py")

            # 2. 删除文件
            client.exec_command("rm -f /opt/ops_agent.py")

            # 3. 更新数据库
            server.use_agent = False
            server.save()

            return f"[{server.hostname}] ✅ 卸载成功"
        except Exception as e:
            return f"[{server.hostname}] ❌ 卸载失败: {str(e)}"
        finally:
            if client: client.close()

    with ThreadPoolExecutor(max_workers=10) as ex:
        futures = [ex.submit(uninstall_one, s) for s in servers]
        results = [f.result() for f in futures]

    return JsonResponse({'status': True, 'log': "\n".join(results)})


@csrf_exempt
def api_receive_metric(request):
    """Agent 上报接口"""
    if request.method != 'POST': return JsonResponse({'status': False}, status=405)
    token = request.headers.get('X-Agent-Token')
    if not token: return JsonResponse({'status': False, 'msg': 'No Token'}, status=403)

    try:
        server = Server.objects.get(agent_token=token)
        d = json.loads(request.body)
        ServerMetric.objects.create(
            server=server,
            cpu_usage=d.get('cpu', 0), mem_usage=d.get('mem', 0),
            disk_usage=d.get('disk', 0), load_1min=d.get('load', 0),
            net_in=d.get('net_in', 0), net_out=d.get('net_out', 0),
            disk_read_rate=d.get('disk_read', 0), disk_write_rate=d.get('disk_write', 0)
        )
        return JsonResponse({'status': True})
    except Server.DoesNotExist:
        return JsonResponse({'status': False}, status=403)
    except Exception as e:
        return JsonResponse({'status': False, 'msg': str(e)}, status=500)


# ========================================================
# 5. Java 应用诊断 (新增)
# ========================================================
# ==========================================
# 1. 页面入口 (加载 Server 和 K8s 集群)
# ==========================================
SIDECAR_IMAGE_JDK8 = "openjdk:8-jdk-alpine"
SIDECAR_IMAGE_JDK17 = "openjdk:17-jdk-alpine"
SIDECAR_IMAGE_ARTHAS = "hengyunabc/arthas:latest"


@login_required
def java_ops_index(request):
    """Java 诊断首页"""
    context = {
        'servers': Server.objects.filter(status='Running'),
        'ai_models': AIModel.objects.all(),
        'page_title': 'Java 应用诊断'
    }

    # 加载 K8s 集群列表
    if K8sCluster:
        context['clusters'] = K8sCluster.objects.all()

    return render(request, 'cmdb/java_diagnose.html', context)


@login_required
@csrf_exempt
def get_java_processes(request):
    """
    SSH 获取 Java 进程列表 (增强版)
    功能：
    1. 扫描宿主机进程
    2. 扫描 Docker 容器
    3. 尝试探测 Java 版本 (Java 8 vs 17)
    """
    sid = request.GET.get('server_id')
    if not sid: return JsonResponse({'status': False})
    server = get_object_or_404(Server, id=sid)

    procs = []
    client = None

    try:
        # 使用安全连接
        client = get_secure_ssh_client(server, timeout=10)

        # === 1. 扫描宿主机 Java 进程 ===
        cmd_host = "ps -eo user,pid,pcpu,pmem,lstart,command | grep java | grep -v grep | head -n 10"
        _, out_host, _ = client.exec_command(cmd_host)
        lines_host = out_host.read().decode().strip().split('\n')

        if lines_host and lines_host[0]:
            for line in lines_host:
                parts = line.split(None, 5)
                if len(parts) >= 6:
                    # 尝试获取宿主机 Java 版本
                    pid = parts[1]
                    ver_cmd = f"java -version 2>&1 | head -n 1"
                    # 这里简化处理：通常宿主机只有一个 Java 版本，或者通过路径推断
                    # 也可以尝试 /proc/{pid}/exe -version，但权限可能不够
                    # 暂且标记为 Host Default
                    java_ver = "Host Default"

                    procs.append({
                        'type': 'host',
                        'container_id': '',
                        'container_name': '-',
                        'user': parts[0],
                        'pid': pid,
                        'cpu': parts[2],
                        'mem': parts[3],
                        'start': parts[4] + " " + parts[5][:4],
                        'cmd': parts[5][:80],
                        'java_version': java_ver
                    })

        # === 2. 扫描 Docker 容器 ===
        # 获取容器 ID, 名称, 镜像
        cmd_docker = "docker ps --format '{{.ID}}|{{.Names}}|{{.Image}}'"
        _, out_docker, _ = client.exec_command(cmd_docker)
        containers = out_docker.read().decode().strip().split('\n')

        for c in containers:
            if not c: continue
            try:
                cid, cname, cimage = c.split('|')

                # 检查是否包含 java 进程 (兼容 Distroless/Alpine/CentOS)
                # 策略：先 exec ps，失败则用 docker top

                # A. 尝试 exec ps (能获取较多信息)
                cmd_check = f"docker exec {cid} ps -ef 2>/dev/null | grep java | grep -v grep | head -n 1"
                _, out_check, _ = client.exec_command(cmd_check)
                res_check = out_check.read().decode().strip()

                target_pid = ""
                cmd_line = ""
                user = "root"

                if res_check:
                    parts = res_check.split(None, 7)
                    if len(parts) >= 8:
                        user = parts[0]
                        target_pid = parts[1]
                        cmd_line = parts[7]
                else:
                    # B. 尝试 docker top (针对 Distroless)
                    cmd_top = f"docker top {cid} -eo pid,comm | grep java"
                    _, out_top, _ = client.exec_command(cmd_top)
                    if out_top.read().decode().strip():
                        target_pid = "1"  # Distroless 中通常是 PID 1
                        cmd_line = "(Distroless/Minimal Image)"

                if target_pid:
                    # === 关键：探测容器内 Java 版本 ===
                    # 只有知道版本，才能在诊断时挂载正确的 JDK 镜像 (8 或 17)
                    ver_check_cmd = f"docker exec {cid} java -version 2>&1 | head -n 1"
                    _, out_ver, _ = client.exec_command(ver_check_cmd)
                    ver_str = out_ver.read().decode().strip()

                    # 解析版本号
                    detected_ver = "Unknown"
                    if "version" in ver_check_cmd or "OpenJDK" in ver_str:
                        if "1.8" in ver_str or '"1.8' in ver_str:
                            detected_ver = "8"
                        elif "11." in ver_str:
                            detected_ver = "11"
                        elif "17." in ver_str:
                            detected_ver = "17"
                        elif "21." in ver_str:
                            detected_ver = "21"
                        else:
                            detected_ver = ver_str[:20]  # 截取一部分

                    procs.append({
                        'type': 'docker',
                        'container_id': cid,
                        'container_name': cname,
                        'user': user,
                        'pid': target_pid,
                        'cpu': 'Docker',
                        'mem': 'Docker',
                        'start': 'N/A',
                        'cmd': f"[{cimage}] {cmd_line}"[:80],
                        'java_version': detected_ver
                    })

            except Exception:
                continue

        return JsonResponse({'status': True, 'data': procs})

    except Exception as e:
        return JsonResponse({'status': False, 'msg': str(e)})
    finally:
        if client: client.close()


# ===========================
# java程序分析
# ===========================
@login_required
@csrf_exempt
def diagnose_java_process(request):
    """
    Java 深度诊断 (Host + Docker Sidecar 模式)
    支持：
    1. 动态选择 JDK8/17 Sidecar 镜像
    2. 处理 Distroless 镜像
    3. 集成 Arthas
    """
    if request.method != 'POST': return JsonResponse({'status': False})

    sid = request.POST.get('server_id')
    pid = request.POST.get('pid')
    mid = request.POST.get('model_id')
    log_path = request.POST.get('log_path', '').strip()

    # 新增参数
    proc_type = request.POST.get('type', 'host')
    container_id = request.POST.get('container_id', '')
    java_version = request.POST.get('java_version', '8')  # 默认为 8

    server = get_object_or_404(Server, id=sid)

    # 确定 Sidecar 镜像
    # 如果检测到是 17, 11, 21 等高版本，使用 JDK17 镜像 (JDK17通常能向下兼容11)
    # 否则默认使用 JDK8
    target_sidecar_image = SIDECAR_IMAGE_JDK8
    if any(v in java_version for v in ['11', '17', '21']):
        target_sidecar_image = SIDECAR_IMAGE_JDK17

    client = None
    try:
        # 使用安全连接
        client = get_secure_ssh_client(server, timeout=20)

        # === 辅助函数：构造命令 ===
        def build_cmd(tool_cmd, tool_type='jdk'):
            """
            tool_type: 'jdk' (jstack/jmap/jstat) | 'arthas' | 'shell' (logs/top)
            """
            if proc_type == 'docker' and container_id:
                # Docker 模式：使用 Sidecar 注入
                if tool_type == 'jdk':
                    # 使用对应的 JDK 镜像，挂载 PID namespace
                    # pid:container 共享进程
                    return f"docker run --rm --pid=container:{container_id} {target_sidecar_image} {tool_cmd}"

                elif tool_type == 'arthas':
                    # 使用 Arthas 镜像，覆盖 entrypoint 执行一次性命令
                    # Arthas 需要 net namespace 才能通信 (虽然 dashboard -n 1 不需要外网，但内部通信需要)
                    return f"docker run --rm --pid=container:{container_id} --net=container:{container_id} --entrypoint /bin/sh {SIDECAR_IMAGE_ARTHAS} -c '{tool_cmd}'"

                elif tool_type == 'shell':
                    # 简单的 Shell 命令，如 logs/stats，直接在宿主机调 docker cli
                    return tool_cmd
            else:
                # Host 模式：直接执行
                if tool_type == 'arthas':
                    # 宿主机 Arthas 逻辑
                    return f"java -jar /tmp/arthas-boot.jar {pid} -c \"{tool_cmd}\" --timeout 20s"
                return tool_cmd

        # === 1. 基础资源 (CPU/MEM) ===
        res = ""
        if proc_type == 'docker':
            # Docker 使用 stats
            cmd_top = f"docker stats {container_id} --no-stream --format 'CPU: {{.CPUPerc}}, MEM: {{.MemUsage}}, NET: {{.NetIO}}'"
            _, out, _ = client.exec_command(cmd_top)
            res = out.read().decode().strip()
        else:
            # Host 使用 top
            cmd_top = f"top -b -n 1 -p {pid} | head -10"
            _, out, _ = client.exec_command(cmd_top)
            res = out.read().decode().strip()

        # === 2. GC 状态 (jstat) ===
        # 注意：PID 1 是指 Sidecar 看到的容器内进程 PID（共享命名空间后通常是 1，但也可能是其他的）
        # 如果是 Host 模式，pid 就是真实的 pid
        target_pid = "1" if proc_type == 'docker' else pid

        cmd_gc = build_cmd(f"jstat -gcutil {target_pid} 1000 2", 'jdk')
        _, out_gc, _ = client.exec_command(cmd_gc)
        gc = out_gc.read().decode().strip()

        # === 3. 内存直方图 (jmap) ===
        cmd_map = build_cmd(f"jmap -histo {target_pid} | head -n 20", 'jdk')
        _, out_map, _ = client.exec_command(cmd_map)
        mem_histo = out_map.read().decode().strip()

        # === 4. JVM 参数 (jinfo) ===
        cmd_info = build_cmd(f"jinfo -flags {target_pid}", 'jdk')
        _, out_info, _ = client.exec_command(cmd_info)
        jvm_flags = out_info.read().decode().strip()

        # === 5. 应用日志 ===
        log_content = ""
        if proc_type == 'docker':
            # Docker 优先看标准输出
            cmd_log = f"docker logs --tail 100 {container_id} 2>&1"
            _, out_log, _ = client.exec_command(cmd_log)
            log_content = out_log.read().decode().strip()
        elif log_path:
            # Host 模式看文件
            cmd_log = f"tail -n 100 {log_path} 2>&1"
            _, out_log, _ = client.exec_command(cmd_log)
            log_content = out_log.read().decode().strip()

        # === 6. Arthas Dashboard ===
        # 准备宿主机 Arthas (仅 Host 模式需要)
        if proc_type == 'host':
            setup_arthas = """
            if [ ! -f /tmp/arthas-boot.jar ]; then
                curl -s -L https://arthas.aliyun.com/arthas-boot.jar -o /tmp/arthas-boot.jar
            fi
            """
            client.exec_command(setup_arthas)

        # 运行 dashboard -n 1
        # Arthas 镜像内 jar 位置通常在 /opt/arthas/arthas-boot.jar
        arthas_run_cmd = f"java -jar /opt/arthas/arthas-boot.jar {target_pid} -c \"dashboard -n 1\" --timeout 20s"
        if proc_type == 'host':
            arthas_run_cmd = "dashboard -n 1"  # build_cmd 里封装了 java -jar ...

        full_arthas_cmd = build_cmd(arthas_run_cmd, 'arthas')
        _, out_arthas, _ = client.exec_command(full_arthas_cmd, timeout=30)
        arthas_dash = out_arthas.read().decode().strip()

        # === 7. 线程堆栈 (jstack) ===
        cmd_stack = build_cmd(f"jstack -l {target_pid}", 'jdk')
        _, out_stack, _ = client.exec_command(cmd_stack)
        stack = out_stack.read().decode().strip()

        # === 构建 Prompt ===
        prompt = f"""
        请扮演一位 Java 性能调优专家，对目标 Java 进程 (PID/Container) 进行全方位诊断。
        环境信息: {proc_type.upper()} Mode, Java Version: {java_version}

        【1. 基础资源】:
        {res}

        【2. GC 状态】:
        {gc}

        【3. 内存对象分布 (Top20)】:
        {mem_histo}

        【4. JVM 启动参数】:
        {jvm_flags}

        【5. Arthas 面板】:
        {arthas_dash}

        【6. 应用日志 (Last 100 lines)】:
        {log_content[:2000]}

        【7. 线程堆栈片段】:
        {stack[:2000]}...

        请输出诊断报告：
        1. **核心结论**：系统当前健康状态（健康/危急/亚健康）。
        2. **风险点分析**：
           - 内存泄漏风险（结合直方图）。
           - GC 频率与耗时。
           - 线程死锁或阻塞。
           - 异常日志归因。
        3. **配置建议**：JVM 参数是否合理？
        4. **优化建议**：给出 3 条具体操作建议。
        """

        ai_res = ask_ai(prompt, model_id=mid, system_role="Java Performance Expert")

        if 'error' in ai_res: return JsonResponse({'status': False, 'msg': ai_res['error']})
        return JsonResponse({'status': True, 'analysis': ai_res['content']})

    except Exception as e:
        return JsonResponse({'status': False, 'msg': f"诊断错误: {str(e)}"})
    finally:
        if client: client.close()


# ===========================
# 6. 云账号管理 (简单增删)
# ===========================
@login_required
def account_list(request):
    if not request.user.is_superuser: return redirect('dashboard')
    if request.method == 'POST':
        form = CloudAccountForm(request.POST)
        if form.is_valid(): form.save()
    return render(request, 'cmdb/account_list.html',
                  {'accounts': CloudAccount.objects.all(), 'form': CloudAccountForm()})


@login_required
def account_delete(request, pk):
    if request.user.is_superuser: CloudAccount.objects.filter(pk=pk).delete()
    return redirect('account_list')




@login_required
@csrf_exempt  # 如果使用了 AJAX POST，建议在前端传递 CSRF Token，或者这里豁免
def server_file_ops(request, server_id):
    """SFTP 文件管理统一接口"""
    server = get_object_or_404(Server, id=server_id)
    action = request.POST.get('action') or request.GET.get('action')
    path = request.POST.get('path') or request.GET.get('path', '/')

    client = None
    sftp = None

    try:
        # 使用安全连接
        client = get_secure_ssh_client(server)
        sftp = client.open_sftp()

        # === 1. 列出目录 (ls) ===
        if action == 'list':
            try:
                # 获取文件属性
                file_list = []
                attrs = sftp.listdir_attr(path)

                for attr in attrs:
                    # 区分文件和目录
                    is_dir = stat.S_ISDIR(attr.st_mode)
                    file_type = 'dir' if is_dir else 'file'

                    # 格式化时间
                    mtime = datetime.fromtimestamp(attr.st_mtime).strftime('%Y-%m-%d %H:%M')

                    # 计算大小 (KB/MB)
                    size_str = ""
                    if not is_dir:
                        if attr.st_size < 1024:
                            size_str = f"{attr.st_size} B"
                        elif attr.st_size < 1024 * 1024:
                            size_str = f"{attr.st_size / 1024:.1f} KB"
                        else:
                            size_str = f"{attr.st_size / (1024 * 1024):.1f} MB"

                    file_list.append({
                        'name': attr.filename,
                        'type': file_type,
                        'size': size_str,
                        'mtime': mtime
                    })

                # 排序：文件夹在前，文件在后
                file_list.sort(key=lambda x: (x['type'] != 'dir', x['name']))

                return JsonResponse({'status': True, 'cwd': path, 'files': file_list})
            except IOError:
                return JsonResponse({'status': False, 'msg': f'无法访问路径: {path}'})

        # === 2. 重命名/移动 (mv) ===
        elif action == 'rename':
            new_path = request.POST.get('new_path')
            try:
                sftp.rename(path, new_path)
                return JsonResponse({'status': True, 'msg': '重命名成功'})
            except Exception as e:
                return JsonResponse({'status': False, 'msg': str(e)})

        # === 3. 删除 (rm) ===
        elif action == 'delete':
            try:
                # 判断是文件还是目录
                try:
                    is_dir = stat.S_ISDIR(sftp.stat(path).st_mode)
                except:
                    return JsonResponse({'status': False, 'msg': '文件不存在'})

                if is_dir:
                    sftp.rmdir(path)  # 注意：rmdir 只能删除空目录，递归删除需要额外写逻辑
                else:
                    sftp.remove(path)
                return JsonResponse({'status': True, 'msg': '删除成功'})
            except Exception as e:
                return JsonResponse({'status': False, 'msg': f'删除失败 (非空目录无法直接删除): {str(e)}'})

        # === 4. 新建文件夹 (mkdir) ===
        elif action == 'mkdir':
            try:
                sftp.mkdir(path)
                return JsonResponse({'status': True, 'msg': '创建成功'})
            except Exception as e:
                return JsonResponse({'status': False, 'msg': str(e)})

        return JsonResponse({'status': False, 'msg': 'Unknown Action'})

    except Exception as e:
        return JsonResponse({'status': False, 'msg': f'SFTP Error: {str(e)}'})
    finally:
        if sftp: sftp.close()
        if client: client.close()


# ========================================================
# 7. Excel 导入导出 (新增)
# ========================================================

@login_required
def server_export(request):
    """导出服务器列表为 Excel"""
    if not request.user.is_superuser:
        return HttpResponse("Permission Denied", status=403)

    # 1. 创建工作簿
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "服务器资产"

    # 2. 写入表头
    headers = ['分组', '主机名', 'IP地址', 'SSH端口', 'SSH用户名', 'SSH密码', 'CPU核数', '内存(GB)', '操作系统', '状态']
    ws.append(headers)

    # 3. 写入数据
    # 根据当前筛选条件导出，或者直接导出全部
    servers = Server.objects.all().select_related('group').order_by('group__name', 'id')

    for s in servers:
        group_name = s.group.name if s.group else "未分组"
        # 注意：s.password 读取时会自动解密
        row = [
            group_name, s.hostname, s.ip_address, s.port,
            s.username, s.password, s.cpu_cores, s.memory_gb,
            s.os_name, s.status
        ]
        ws.append(row)

    # 4. 返回响应
    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = f'attachment; filename="server_list_{datetime.now().strftime("%Y%m%d")}.xlsx"'
    wb.save(response)
    return response


@login_required
@csrf_exempt
def server_import(request):
    """从 Excel 导入服务器"""
    if not request.user.is_superuser:
        return JsonResponse({'status': False, 'msg': '权限不足'})

    if request.method != 'POST':
        return JsonResponse({'status': False, 'msg': '方法不允许'})

    file = request.FILES.get('file')
    if not file:
        return JsonResponse({'status': False, 'msg': '请选择文件'})

    try:
        wb = openpyxl.load_workbook(file)
        ws = wb.active

        success_count = 0
        error_rows = []

        # 从第2行开始读取 (跳过表头)
        for i, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
            try:
                # 对应导出时的列顺序
                # 0:分组, 1:主机名, 2:IP, 3:端口, 4:用户, 5:密码, 6:CPU, 7:内存, 8:OS, 9:状态
                group_name, hostname, ip, port, user, pwd, cpu, mem, os_name, status = row[:10]

                if not ip or not hostname: continue  # 跳过空行

                # 处理分组 (不存在则自动创建)
                group = None
                if group_name and group_name != "未分组":
                    group, _ = ServerGroup.objects.get_or_create(name=group_name)

                # 更新或创建服务器 (按IP排重)
                Server.objects.update_or_create(
                    ip_address=ip,
                    defaults={
                        'hostname': hostname,
                        'group': group,
                        'port': int(port) if port else 22,
                        'username': user or 'root',
                        'password': pwd or '',
                        'cpu_cores': int(cpu) if cpu else 1,
                        'memory_gb': int(mem) if mem else 1,
                        'os_name': os_name or 'Linux',
                        'status': status or 'Running',
                        'provider': 'private'  # 默认标记为私有
                    }
                )
                success_count += 1
            except Exception as e:
                error_rows.append(f"第 {i} 行 ({row[2] if len(row) > 2 else '未知IP'}): {str(e)}")

        msg = f"导入完成！成功: {success_count} 条。"
        if error_rows:
            msg += f"\n失败 {len(error_rows)} 条:\n" + "\n".join(error_rows)

        return JsonResponse({'status': True, 'msg': msg})

    except Exception as e:
        return JsonResponse({'status': False, 'msg': f"文件解析失败: {str(e)}"})


@login_required
def ssl_cert_list(request):
    """SSL 证书列表页"""
    certs = SSLCertificate.objects.all().order_by('remaining_days')

    # 获取 Webhook 配置用于回显到模态框
    ding_conf = SystemConfig.objects.filter(key='dingtalk_webhook').first()
    wechat_conf = SystemConfig.objects.filter(key='wechat_webhook').first()

    context = {
        'certs': certs,
        'dingtalk_webhook': ding_conf.value if ding_conf else '',
        'wechat_webhook': wechat_conf.value if wechat_conf else ''
    }
    return render(request, 'cmdb/ssl_cert_list.html', context)


@login_required
@csrf_exempt
def ssl_cert_add(request):
    """[AJAX] 添加域名"""
    if request.method == 'POST':
        domain = request.POST.get('domain')
        port = request.POST.get('port', 443)
        if not domain:
            return JsonResponse({'status': False, 'msg': '请输入域名'})

        obj, created = SSLCertificate.objects.get_or_create(domain=domain, defaults={'port': port})
        if not created:
            return JsonResponse({'status': False, 'msg': '域名已存在'})

        # 添加后立即触发一次异步检测
        check_ssl_certificates_task.delay()
        return JsonResponse({'status': True, 'msg': '添加成功，正在后台检测...'})


@login_required
def ssl_cert_delete(request, pk):
    """删除域名"""
    SSLCertificate.objects.filter(id=pk).delete()
    return redirect('ssl_cert_list')


@login_required
def ssl_cert_refresh(request):
    """手动触发全量刷新"""
    check_ssl_certificates_task.delay()
    messages.success(request, "已触发后台刷新任务，请稍后刷新页面查看结果")
    return redirect('ssl_cert_list')


@login_required
@csrf_exempt
def ssl_config_save(request):
    """[AJAX] 保存告警 Webhook 配置"""
    if request.method == 'POST':
        ding = request.POST.get('dingtalk_webhook', '').strip()
        wechat = request.POST.get('wechat_webhook', '').strip()

        SystemConfig.objects.update_or_create(key='dingtalk_webhook', defaults={'value': ding})
        SystemConfig.objects.update_or_create(key='wechat_webhook', defaults={'value': wechat})

        return JsonResponse({'status': True, 'msg': '配置已保存'})